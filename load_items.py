"""
load_items.py — PART 1: item master loader (clears Blocker 1 — cost prices + GST mapping).

Reads the filled Excel template K24_Zoho_Item_Import_Template.xlsx (sheet '② Item List'),
VALIDATES every row, then creates/upserts Zoho Inventory items — idempotent on SKU.

SAFE-MODE contract (the ONLY go-live switch is LIVE_MODE):
  * no --commit                      -> full DRY RUN: validate + print the exact would-do
                                        table, write NOTHING.
  * --commit while LIVE_MODE=false   -> write only the first N (default 2) items as a TEST
                                        BATCH, then STOP for human confirmation.
  * --commit while LIVE_MODE=true    -> full load (idempotent).

Validation STOPS before any API call if a REQUIRED field is missing: Item Name, SKU, Type,
Unit, HSN, GST Rate, Selling Price, Cost Price, Opening Stock (and Opening Stock Rate if
Opening Stock > 0). Also flags: duplicate SKUs, non-numeric prices, GST not in
{0,5,12,18,28}, HSN not 8 digits.

GST is mapped to the org's EXISTING tax for that rate — a missing tax is reported, never
invented. HSN/GST come only from the template — blanks are flagged, never guessed.

Outputs: item_load_report.md + item_load_raw.json.

Run:
  python load_items.py                         # dry run (default file)
  python load_items.py --file other.xlsx       # dry run of another file
  python load_items.py --commit                # test batch (first 2) if LIVE_MODE=false
  python load_items.py --commit --limit 5      # test batch of 5
  python load_items.py --commit --full         # full load (allowed only if LIVE_MODE=true)
"""
from __future__ import annotations

import argparse
import asyncio
import json
import os
from typing import Any

import structlog

import golive_common as gc
from zoho_client import TOKEN_REGEN_RUNBOOK, ZohoAuthError, ZohoClient, ZohoError

log = structlog.get_logger("load_items")

DEFAULT_FILE = "K24_Zoho_Item_Import_Template.xlsx"
SHEET_HINT = "item list"  # matches '② Item List' regardless of the leading glyph
REPORT_MD = "item_load_report.md"
REPORT_JSON = "item_load_raw.json"

REQUIRED = ["Item Name", "SKU", "Type", "Unit", "HSN Code", "GST Rate (%)", "Selling Price", "Cost Price", "Opening Stock"]

# canonical header -> accepted variants (lowercased, stripped)
COLUMN_ALIASES = {
    "Item Name": {"item name", "name"},
    "SKU": {"sku"},
    "Type": {"type"},
    "Unit": {"unit", "uom"},
    "HSN Code": {"hsn code", "hsn", "hsn/sac", "hsn_or_sac"},
    "GST Rate (%)": {"gst rate (%)", "gst rate", "gst", "gst%", "tax rate"},
    "Selling Price": {"selling price", "rate", "sell price", "mrp"},
    "Cost Price": {"cost price", "cost", "purchase rate", "purchase price"},
    "Opening Stock": {"opening stock", "opening", "opening qty"},
    "Opening Stock Rate": {"opening stock rate", "opening rate", "opening stock value rate"},
    "Reorder Level": {"reorder level", "reorder", "reorder point"},
    "Category": {"category", "item group", "group"},
    "Description": {"description", "desc"},
}


# ------------------------------------------------------------------ read xlsx
def _norm(s: Any) -> str:
    return str(s or "").strip().lower()


def read_rows(path: str) -> list[dict[str, Any]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = next((s for s in wb.sheetnames if SHEET_HINT in s.lower()), None)
    if sheet is None:
        wb.close()
        raise ValueError(f"No '② Item List' sheet in {path}. Sheets present: {wb.sheetnames}")
    ws = wb[sheet]

    rows_iter = ws.iter_rows(values_only=True)
    # find the header row (the first row that contains an 'item name'/'sku' header)
    header: list[str] | None = None
    header_map: dict[int, str] = {}
    for raw in rows_iter:
        cells = [_norm(c) for c in raw]
        if any(c in COLUMN_ALIASES["Item Name"] for c in cells) and any(c in COLUMN_ALIASES["SKU"] for c in cells):
            header = list(raw)
            for idx, c in enumerate(cells):
                for canon, variants in COLUMN_ALIASES.items():
                    if c in variants:
                        header_map[idx] = canon
                        break
            break
    if header is None:
        wb.close()
        raise ValueError(f"Could not locate a header row (needs 'Item Name' + 'SKU') in sheet '{sheet}' of {path}.")

    out: list[dict[str, Any]] = []
    for excel_rownum, raw in enumerate(rows_iter, start=2):  # data begins after header
        if raw is None or all(c is None or str(c).strip() == "" for c in raw):
            continue  # skip blank rows
        rec: dict[str, Any] = {"_row": excel_rownum}
        for idx, val in enumerate(raw):
            canon = header_map.get(idx)
            if canon:
                rec[canon] = val
        # ignore rows with no SKU and no name (stray formatting)
        if not str(rec.get("SKU", "")).strip() and not str(rec.get("Item Name", "")).strip():
            continue
        out.append(rec)
    wb.close()
    return out


# ------------------------------------------------------------------ validate
def _to_num(v: Any) -> float | None:
    if v is None or str(v).strip() == "":
        return None
    try:
        return float(str(v).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def validate(rows: list[dict[str, Any]]) -> tuple[list[dict], list[str]]:
    """Return (validated_rows_with_parsed_fields, errors). errors non-empty => STOP."""
    errors: list[str] = []
    seen_sku: dict[str, int] = {}
    valid: list[dict] = []

    for r in rows:
        rownum = r["_row"]
        rerrs: list[str] = []

        name = str(r.get("Item Name", "")).strip()
        sku = str(r.get("SKU", "")).strip()
        typ = str(r.get("Type", "")).strip()
        unit = str(r.get("Unit", "")).strip()
        hsn = str(r.get("HSN Code", "")).strip()
        # HSN read from Excel can arrive as 21069099.0
        if hsn.endswith(".0"):
            hsn = hsn[:-2]
        gst = _to_num(r.get("GST Rate (%)"))
        selling = _to_num(r.get("Selling Price"))
        cost = _to_num(r.get("Cost Price"))
        opening = _to_num(r.get("Opening Stock"))
        opening_rate = _to_num(r.get("Opening Stock Rate"))
        reorder = _to_num(r.get("Reorder Level"))
        category = str(r.get("Category", "")).strip()
        desc = str(r.get("Description", "")).strip()

        # required present
        for label, present in [
            ("Item Name", bool(name)), ("SKU", bool(sku)), ("Type", bool(typ)), ("Unit", bool(unit)),
            ("HSN Code", bool(hsn)), ("GST Rate (%)", gst is not None),
            ("Selling Price", selling is not None), ("Cost Price", cost is not None),
            ("Opening Stock", opening is not None),
        ]:
            if not present:
                rerrs.append(f"missing required '{label}'")
        if opening is not None and opening > 0 and opening_rate is None:
            rerrs.append("Opening Stock > 0 but 'Opening Stock Rate' is blank")

        # flags
        if sku:
            if sku in seen_sku:
                rerrs.append(f"duplicate SKU (also row {seen_sku[sku]})")
            else:
                seen_sku[sku] = rownum
        if r.get("Selling Price") not in (None, "") and selling is None:
            rerrs.append("Selling Price is non-numeric")
        if r.get("Cost Price") not in (None, "") and cost is None:
            rerrs.append("Cost Price is non-numeric")
        if gst is not None and int(gst) not in gc.VALID_GST_RATES:
            rerrs.append(f"GST rate {gst} not in {sorted(gc.VALID_GST_RATES)}")
        if hsn and (not hsn.isdigit() or len(hsn) != 8):
            rerrs.append(f"HSN '{hsn}' is not 8 digits")

        if rerrs:
            errors.extend(f"row {rownum} ({sku or name or '?'}): {e}" for e in rerrs)
            continue

        valid.append({
            "_row": rownum, "name": name, "sku": sku, "type": typ, "unit": unit, "hsn": hsn,
            "gst": int(gst), "selling": selling, "cost": cost, "opening": opening,
            "opening_rate": opening_rate, "reorder": reorder, "category": category, "desc": desc,
        })
    return valid, errors


# ------------------------------------------------------------------ tax mapping
def resolve_tax(rate: int, tax_map: dict[int, dict]) -> tuple[dict, str | None]:
    """Return (item tax fields, error). Never invents a tax id."""
    if rate == 0:
        # 0% GST — non-taxable line. Use a 0% tax if the org has one, else mark non-taxable.
        if 0 in tax_map:
            return {"is_taxable": True, "tax_id": tax_map[0]["tax_id"]}, None
        return {"is_taxable": False, "tax_exemption_code": "NONTAXABLE"}, None
    t = tax_map.get(rate)
    if not t:
        return {}, f"no org tax configured for GST {rate}% — create it in Settings -> Taxes (not invented)"
    return {"is_taxable": True, "tax_id": t["tax_id"]}, None


# ------------------------------------------------------------------ payload
def build_create_payload(row: dict, tax_fields: dict, category_id: str | None) -> dict:
    payload: dict[str, Any] = {
        "name": row["name"],
        "sku": row["sku"],
        "product_type": "goods",
        "item_type": "inventory",
        "unit": row["unit"],
        "hsn_or_sac": row["hsn"],
        "rate": row["selling"],
        "purchase_rate": row["cost"],
        "track_inventory": True,
    }
    if row.get("reorder") is not None:
        payload["reorder_level"] = row["reorder"]
    if row.get("desc"):
        payload["description"] = row["desc"]
    if category_id:
        payload["category_id"] = category_id
    payload.update(tax_fields)
    # opening stock at the target location (multi-location org)
    if row.get("opening") and row["opening"] > 0:
        rate = row.get("opening_rate") or row["cost"]
        payload["opening_stock"] = row["opening"]
        payload["opening_stock_value"] = round(row["opening"] * rate, 2)
        if gc.LOCATION_ID:
            payload["locations"] = [{
                "location_id": gc.LOCATION_ID,
                "initial_stock": row["opening"],
                "initial_stock_rate": rate,
            }]
    return payload


def build_update_payload(row: dict, tax_fields: dict, category_id: str | None) -> dict:
    # Opening stock is NOT updated for existing items (stock already tracked) — cost/price/GST only.
    payload: dict[str, Any] = {
        "name": row["name"],
        "unit": row["unit"],
        "hsn_or_sac": row["hsn"],
        "rate": row["selling"],
        "purchase_rate": row["cost"],
    }
    if row.get("reorder") is not None:
        payload["reorder_level"] = row["reorder"]
    if row.get("desc"):
        payload["description"] = row["desc"]
    if category_id:
        payload["category_id"] = category_id
    payload.update(tax_fields)
    return payload


# ------------------------------------------------------------------ main
async def run(path: str, commit: bool, limit: int, full: bool) -> dict:
    result: dict[str, Any] = {"file": path, "committed": commit, "planned": [], "created": [], "updated": [], "skipped": [], "errors": []}

    if not os.path.exists(path):
        msg = (f"Template file not found: {path}. Place the filled "
               f"'K24_Zoho_Item_Import_Template.xlsx' (② Item List sheet) in {os.getcwd()} "
               "or pass --file <path>.")
        print("STOP:", msg)
        result["errors"].append(msg)
        _write_report(result, validation_errors=[msg], mode="stopped")
        return result

    rows = read_rows(path)
    print(f"Read {len(rows)} data rows from '{path}'.")
    valid, verrs = validate(rows)

    if verrs:
        print(f"\nVALIDATION FAILED — {len(verrs)} problem(s). No API calls made:")
        for e in verrs:
            print("  -", e)
        result["errors"] = verrs
        _write_report(result, validation_errors=verrs, valid_rows=valid, mode="validation_failed")
        return result
    print(f"Validation passed: {len(valid)} rows OK.")

    # -- connect, load org taxes + categories (read-only planning) ---------
    try:
        async with ZohoClient() as z:
            taxes = await gc.fetch_taxes(z)
            tax_map = gc.build_tax_rate_map(taxes)
            cats = (await z.get(z.inventory("/categories"))).get("categories", []) or []
            cat_by_name = {_norm(c.get("name")): c.get("category_id") for c in cats}

            # build a plan (also catches tax-mapping gaps before any write)
            plan: list[dict] = []
            for row in valid:
                tax_fields, tax_err = resolve_tax(row["gst"], tax_map)
                cat_id = cat_by_name.get(_norm(row["category"])) if row["category"] else None
                cat_note = None
                if row["category"] and not cat_id:
                    cat_note = f"category '{row['category']}' not found in org — item will load without a category"
                plan.append({"row": row, "tax_fields": tax_fields, "tax_err": tax_err, "cat_id": cat_id, "cat_note": cat_note})

            tax_gaps = [f"row {p['row']['_row']} {p['row']['sku']}: {p['tax_err']}" for p in plan if p["tax_err"]]
            if tax_gaps:
                print(f"\nGST MAPPING GAPS — {len(tax_gaps)} row(s) reference a tax the org doesn't have:")
                for g in tax_gaps:
                    print("  -", g)
                print("  (Fix: create these tax rates in the UI/CA step, then re-run. Tax ids are never invented.)")
            result["tax_rates_available"] = sorted(tax_map.keys())

            # -- decide what actually gets written --------------------------
            if not commit:
                mode = "dry_run"
                to_write: list[dict] = []
            elif gc.LIVE_MODE and full:
                mode = "commit_full_live"
                to_write = plan
            elif gc.LIVE_MODE and not full:
                mode = "commit_test_batch_live"
                to_write = plan[:limit]
            else:  # commit while LIVE_MODE=false -> test batch only, then stop
                mode = "commit_test_batch_safe"
                to_write = plan[:limit]

            # -- resolve existing items for idempotency (planning display) --
            for p in plan:
                existing = await gc.find_item_by_sku(z, p["row"]["sku"])
                p["exists"] = bool(existing)
                p["existing_id"] = existing.get("item_id") if existing else None
                action = "update" if existing else "create"
                if p["tax_err"]:
                    action = "SKIP (tax gap)"
                result["planned"].append({
                    "sku": p["row"]["sku"], "name": p["row"]["name"], "action": action,
                    "cost": p["row"]["cost"], "gst": p["row"]["gst"], "hsn": p["row"]["hsn"],
                    "opening": p["row"]["opening"], "tax_err": p["tax_err"], "cat_note": p["cat_note"],
                })

            _print_plan(result["planned"], mode, len(to_write))

            # -- write (only when committing) -------------------------------
            if commit:
                if gc.LIVE_MODE and full:
                    print(f"\nLIVE_MODE=true + --full -> writing ALL {len(to_write)} items.")
                else:
                    print(f"\nWriting TEST BATCH of {len(to_write)} item(s) "
                          f"(LIVE_MODE={'true' if gc.LIVE_MODE else 'false'}). Will STOP after for confirmation.")
                for p in to_write:
                    if p["tax_err"]:
                        result["skipped"].append({"sku": p["row"]["sku"], "reason": p["tax_err"]})
                        continue
                    try:
                        if p["exists"]:
                            payload = build_update_payload(p["row"], p["tax_fields"], p["cat_id"])
                            res = (await z.put(z.inventory(f"/items/{p['existing_id']}"), json=payload)).get("item", {})
                            result["updated"].append({"sku": p["row"]["sku"], "item_id": res.get("item_id") or p["existing_id"]})
                            print(f"  updated {p['row']['sku']} ({res.get('item_id') or p['existing_id']})")
                        else:
                            payload = build_create_payload(p["row"], p["tax_fields"], p["cat_id"])
                            res = (await z.post(z.inventory("/items"), json=payload)).get("item", {})
                            result["created"].append({"sku": p["row"]["sku"], "item_id": res.get("item_id")})
                            print(f"  created {p['row']['sku']} ({res.get('item_id')})")
                    except ZohoError as e:
                        result["errors"].append({"sku": p["row"]["sku"], "error": str(e), "payload": getattr(e, "payload", None)})
                        print(f"  ERROR {p['row']['sku']}: {e}")

                if not (gc.LIVE_MODE and full):
                    print("\nSTOP: test batch complete. Review item_load_report.md, then re-run with "
                          "'--commit --full' AFTER LIVE_MODE=true to load the rest.")

            _write_report(result, validation_errors=[], valid_rows=valid, mode=mode, tax_gaps=tax_gaps)
            return result

    except ZohoAuthError as e:
        print("\nAUTH FAILED:", e)
        print(TOKEN_REGEN_RUNBOOK)
        result["errors"].append({"auth": str(e)})
        _write_report(result, validation_errors=[], valid_rows=valid, mode="auth_failed")
        return result


def _print_plan(planned: list[dict], mode: str, n_write: int) -> None:
    print(f"\n=== PLAN ({mode}) — {len(planned)} items; {n_write} would be written this run ===")
    rows = [[p["action"], p["sku"], (p["name"] or "")[:34], p["cost"], f"{p['gst']}%", p["hsn"], p["opening"],
             (p["tax_err"] or p["cat_note"] or "")] for p in planned]
    print(gc.md_table(["Action", "SKU", "Name", "Cost", "GST", "HSN", "OpenStk", "Note"], rows))


def _write_report(result: dict, *, validation_errors: list[str], valid_rows: list[dict] | None = None,
                  mode: str = "", tax_gaps: list[str] | None = None) -> None:
    lines = [
        "# Item Master Load Report (Blocker 1)",
        "",
        f"- **File:** {result.get('file')}",
        f"- **Mode:** {mode}",
        f"- **LIVE_MODE:** {'true' if gc.LIVE_MODE else 'false'}",
        f"- **Target location (opening stock):** {gc.LOCATION_ID or '(ZOHO_LOCATION_ID unset)'}",
        "",
    ]
    if validation_errors:
        lines += ["## Validation errors (STOPPED — no API calls)", *[f"- {e}" for e in validation_errors], ""]
    if tax_gaps:
        lines += ["## GST mapping gaps (tax id never invented)", *[f"- {g}" for g in tax_gaps], ""]
    if result.get("planned"):
        lines += ["## Plan", gc.md_table(
            ["Action", "SKU", "Name", "Cost", "GST", "HSN", "OpenStk"],
            [[p["action"], p["sku"], p["name"], p["cost"], f"{p['gst']}%", p["hsn"], p["opening"]] for p in result["planned"]]), ""]
    for label, key in [("Created", "created"), ("Updated", "updated"), ("Skipped", "skipped"), ("Errors", "errors")]:
        if result.get(key):
            lines += [f"## {label} ({len(result[key])})", "```json", json.dumps(result[key], indent=2, default=str), "```", ""]
    with open(REPORT_MD, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    with open(REPORT_JSON, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, default=str)
    print(f"\nWrote {REPORT_MD} and {REPORT_JSON}")


def _parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(description="Load K24/Kingdom Foods items into Zoho Inventory (safe-mode).")
    ap.add_argument("--file", default=DEFAULT_FILE, help=f"Excel template (default {DEFAULT_FILE})")
    ap.add_argument("--commit", action="store_true", help="actually write (test batch unless --full + LIVE_MODE=true)")
    ap.add_argument("--limit", type=int, default=2, help="test-batch size (default 2)")
    ap.add_argument("--full", action="store_true", help="full load — only honored when LIVE_MODE=true")
    return ap.parse_args()


if __name__ == "__main__":
    args = _parse_args()
    if args.full and not gc.LIVE_MODE:
        print("NOTE: --full ignored while LIVE_MODE=false. Writing a test batch only (safe-mode guard).")
        args.full = False
    asyncio.run(run(args.file, args.commit, args.limit, args.full))
